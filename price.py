import json
import shutil
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import requests
import xml.etree.ElementTree as ET
# import win32com.client as win32
from yandex import EmailDownloader
import math
from datetime import timedelta, date
from pprint import pprint

AUTOFAMILY_API_KEY = "mMS6OL8RtP5Ca9iHCfvAIGUz5Qcd39fB"
COLORS = {
    "RED": "F2DCD7",
    "YELLOW": "f7f5dd",
    "light_gray_color": "FFBFBFBF",
    "normal_dark_gray_color": "FF808080",
    "medium_dark_gray_color": "FF606060",
    "strong_dark_gray_color": "FF404040",
}


def read_int(value):
    return int(value) if value is not None else 0


def pixels_to_width(pixels):
    return (pixels - 5) / 7


# Загружаем XML-файл
def download_xml(url, save_path):
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()  # Проверка на ошибки HTTP
        with open(save_path, 'wb') as file:
            for chunk in response.iter_content(chunk_size=8192):
                file.write(chunk)
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при загрузке файла: {e}")


def get_warehouses_data(article, api_key):
    """
    Делает запрос к API, извлекает и возвращает данные из ключа 'warehouses'.

    :param article: Артикул продукта для запроса.
    :param api_key: API-ключ пользователя.
    :return: Словарь с данными 'warehouses' или пустой словарь, если ключ отсутствует.
    """
    url = f"https://b2b.autofamily.ru/api/getProduct/?article={
        article}&key={api_key}&v=1.1"
    try:
        # Отправка GET-запроса
        response = requests.get(url)
        response.raise_for_status()  # Проверка на успешность ответа
        print(response)

        # Разбор JSON-ответа
        data = response.json()

        # Извлечение ключа 'response'
        response_data = data.get("response", {})
        pprint(response_data)

        name = response_data.get("name", "")

        # Проверка наличия ключа 'warehouses' в 'response'
        warehouses = response_data.get("warehouses", {})
        if not isinstance(warehouses, dict):
            raise ValueError("'warehouses' должен быть словарём")

        result = {"warehouses": warehouses, "name": name}

        return result

    except requests.exceptions.RequestException as e:
        print(f"Ошибка соединения с API: {e}")
    except ValueError as e:
        print(f"Ошибка обработки данных: {e}")
    except Exception as e:
        print(f"Непредвиденная ошибка: {e}")

    # Возвращаем пустой словарь, если произошла ошибка
    return {}


# def download_xml(url, save_path):
#     try:
#         urllib.request.urlretrieve(url, save_path)
#     except Exception as e:
#         print(f"Ошибка при загрузке файла: {e}")


# Парсим XML-файл
def parse_xml(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Создаем словарь для хранения соответствий ARTICLE и WEIGHT
    article_weight = {}
    for item in root.findall('.//ITEM'):
        article = item.find('ARTICLE').text
        article = article.upper()
        weight_str = item.find('WEIGHT').text
        weight = float(weight_str)
        weight = math.ceil(weight * 100) / 100
        article_weight[article] = weight
    return article_weight


class AvitoFilesHandler:
    def __init__(self, settings_file):
        self.settings_file = settings_file
        self.settings = self.load_settings(settings_file)

    def load_settings(self, settings_file):
        with open(settings_file, 'r', encoding='utf-8') as file:
            return json.load(file)  # Возвращаем весь содержимое файла

    def load_ostatki(self):
        email_downloader = EmailDownloader(self.settings_file)
        email_downloader.connect()
        email_downloader.fetch_emails()
        email_downloader.logout()

    
    def read_ostatki(self):
        ostatki_path = self.settings["Настройки"]["Остатки"]

        # Считываем данные из Excel "Остатки"
        price = self.excel_to_list(ostatki_path)
        price.pop(0)

        # Получаем настройки аббревиатур городов
        cities = self.settings["Аббревиатуры городов"]
        # pprint(cities)

        # Ключи словаря данных
        headers = list(price[0].keys())

        # Дообрабатываем считанные данные 
        for item in price:
            item["Вес"] = None
            item["Link1"] = ""
            item["Link2"] = ""
            item["Link3"] = ""
            item["ID_01"] = ""
            item["Avitoid_01"] = ""
            item["Avitoid_02"] = ""
            item["Avitoid_03"] = ""
            item["api_json"] = ""
            item["Прайс"] = None
            item["Цена"] = item.pop("Цена без распродажи")
            item["Расп"] = item.pop("Цена распродажи")
            if item["Цена"]:
                item["д10%"] = round(item["Цена"] * 0.9, 2)
            # Замена наименований городов на аббревиатуры 
            for city in cities:
                for header in headers:
                    if header == city:
                        item[cities[city]["Аббревиатура"]] = item.pop(header)
            item["AvitoStatus"] = ""
            if item["Марка"] in [None, 0]:
                item["Марка"] = ""
            if item["Модель"] in [None, 0]:
                item["Модель"] = ""
            if item["Бренд"] in [None, 0]:
                item["Бренд"] = ""
            if item["Товарная группа"] in [None, 0]:
                item["Товарная группа"] = ""
            
        # for i in range(0, 2):
        #     pprint(price[i])
        return price

    def update_price_from_products(self, price):
        # price = self.read_ostatki()

        goods_path = self.settings["Настройки"]["Таблица_с_товарами"]
        link_0 = self.settings["Ссылки"]["Ссылка_прайса"]

        # Загружаем таблицы
        goods_workbook = load_workbook(goods_path)
        goods_sheet = goods_workbook.active

        # Находим индекс столбца "Артикул" в таблице "Товар"
        article_col_goods = None
        for col in range(1, goods_sheet.max_column + 1):
            if goods_sheet.cell(row=1, column=col).value == "Артикул":
                article_col_goods = col
                break

        if article_col_goods is None:
            print("Не удалось найти столбец 'Артикул' в таблице 'Товары'.")
            return

        avito_settings = self.settings["Авито"]  # Обращаемся к разделу "Авито"
        avito_count = len(avito_settings)

        # Создаем словарь для значений из таблицы "Товар"
        goods_data = self.excel_to_dict(
            excel_path=self.settings["Настройки"]["Таблица_с_товарами"], key_col_header="Артикул", key_upper=True)
        for key in goods_data:
            goods_data[key]["is_correct"] = False
            if goods_data[key]["Артикул"] is not None:
                goods_data[key]["Артикул"] = goods_data[key]["Артикул"].upper()
                for code in range(1, avito_count + 1):
                    if goods_data[key][f"{code:02}: AvitoId"] != "":
                        goods_data[key]["is_correct"] = True

        # Скачиваем XML-файл
        print('Начало скачивания файла XML')
        xml_path = self.settings["Настройки"]["XML_файл"]
        download_xml(
            'http://portal.autofamily.ru/export/autofamily_catalog.xml', xml_path)
        print('Скачивание файла XML завершено')

        # Парсим XML-файл
        article_weight = parse_xml(xml_path)

        # Вставляем веса и значения из таблицы "Товар" в таблицу "Прайс"
        for item in price:
            article_value = item["Артикул"].upper()
            # Вставляем веса
            if article_value in article_weight:
                weight = article_weight[article_value]
                # Записываем вес в 1-й столбец
                item["Вес"] = weight
            # Вставляем значения из таблицы "Товар"
            if (article_value in goods_data) and goods_data[article_value]["is_correct"]:
                item["ID_01"] = goods_data[article_value]["01: Id"]

                for code in range(1, avito_count + 1):
                    item[f"Avitoid_{code:02}"] = goods_data[article_value][f"{code:02}: AvitoId"]

                item["Прайс"] = goods_data[article_value]["01: Price"]
                item["AvitoStatus"] = goods_data[article_value]["AvitoStatus"]

                for code in range(1, avito_count + 1):
                    item[f"Link{code}"] = link_0.replace("{ID_XX}", str(goods_data[article_value][f"{code:02}: AvitoId"]))

        # self.apply_filters_to_price_sheet(price_sheet, article_col_price, brand_col_price, product_group_col_price)

        print(f"Данные 'Прайс' обновлены на основании таблицы 'Товар'.")
        return price


    def apply_filters_to_price(self, price):
        # Получаем фильтры из настроек
        text_filters = self.settings["Фильтры"]["Фильтр_по_тексту"]
        article_filters = self.settings["Фильтры"]["Фильтр_по_артикулу"]
        group_brand_filters = self.settings["Фильтры"]["Фильтр_по_товарная_группа_и_бренд"]

        # Удаляем вхождения текстовых фильтров из Товаров
        for i in range(0, len(price)):  # Начинаем с 2-й строки
            product = price[i]["Товар"]
            if product:  # Проверяем, что значение не None
                for text_filter in text_filters:
                    product = product.replace(text_filter, "")  # Заменяем на пустую строку
                # Убираем лишние пробелы
                price[i]["Товар"] = product.strip()

        # Помечаем на удаление строки, если найдены соответствия по Артикулу, по Бренду или по Группе товаров
        rows_to_delete = []
        for i in range(0, len(price)):
            if ((price[i]["Артикул"] in [None, '', 0])
                or (price[i]["Артикул"] in article_filters) 
                or (price[i]["Бренд"] in group_brand_filters) 
                or (price[i]["Товарная группа"] in group_brand_filters)):
                rows_to_delete.append(i)

        # Удаляем строки, начиная с конца, чтобы не нарушить индексацию
        for i in reversed(rows_to_delete):
            price.pop(i)

        print(f"Для 'Прайс' применены фильтры.")

        return price
    

    def sort_price(self, price):
        # Получаем фильтры из настроек
        pg_order = self.settings["Сортировка"]["Сортировка по товарная группа"]
        brand_order = self.settings["Сортировка"]["Сортировка по бренд"]

        # Функция сортировки
        def sort_by_priority(product):
            # Получаем приоритет для позиций сортировки
            marka_value = product["Марка"] if product["Марка"] else ""
            model_value = product["Модель"] if product["Модель"] else ""
            mudguards_priority = pg_order.index("Брызговики") if "Брызговики" in pg_order else float("inf")
            if product["Товарная группа"].startswith("Брызговики"):
                pg_priority = mudguards_priority
                brand_priority = float("inf")
                pg_value = "Брызговики"
                brand_value = "Брызговики"
            else:
                pg_priority = pg_order.index(product["Товарная группа"]) if product["Товарная группа"] in pg_order else float("inf")
                brand_priority = brand_order.index(product["Бренд"]) if product["Бренд"] in brand_order else float("inf")
                pg_value = product["Товарная группа"] if product["Товарная группа"] else ""
                brand_value = product["Бренд"] if product["Бренд"] else ""
            # Возвращаем кортеж для сортировки
            return (marka_value, model_value, pg_priority, pg_value, brand_priority, brand_value)

        sorted_price = sorted(price, key=sort_by_priority)

        print(f"'Прайс' отсортирован.")

        return sorted_price
    

    def create_price_xlsx(self):
        hidden_cols = ["ID_01","Avitoid_01","Avitoid_02","Avitoid_03","д10%"]
        cell_border_color = self.settings["Цвет"]["Граница товара"]
        pg_border_color = self.settings["Цвет"]["Граница товарной группы"]
        model_border_color = self.settings["Цвет"]["Граница модели"]
        marka_border_color = self.settings["Цвет"]["Граница марки"]

        # Настраиваем стандартный цвет линий сетки (визуализация через границы)
        thin_border = Border(
            top=Side(border_style="thin", color=cell_border_color),  
            left=Side(border_style="thin", color=cell_border_color),
            right=Side(border_style="thin", color=cell_border_color),
            bottom=Side(border_style="thin", color=cell_border_color),
        )

        pg_border = Border(
            top=Side(border_style="thin", color=pg_border_color), 
            left=thin_border.left,
            right=thin_border.right,
            bottom=thin_border.bottom,
        )
        
        model_border = Border(
            top=Side(border_style="medium", color=model_border_color),  
            left=thin_border.left,
            right=thin_border.right,
            bottom=thin_border.bottom,
        )

        marka_border = Border(
            top=Side(border_style="thick", color=marka_border_color), 
            left=thin_border.left,
            right=thin_border.right,
            bottom=thin_border.bottom,
        )

        print(f"Начата вставка данных в 'Прайс.xlsm'")
        price = self.read_ostatki()
        price = self.update_price_from_products(price)
        price = self.apply_filters_to_price(price)
        price = self.sort_price(price)
        # pprint(orders_data)

        # Формируем Excel файлы
        price_template = self.settings["Настройки"]["Шаблон_для_Прайса"]
        price_file_path = self.settings["Настройки"]["Прайс_для_работы"]
        shutil.copy(price_template, price_file_path)

 
        # Открываем скопированный файл
        wb = load_workbook(price_file_path, keep_vba=True)
        sheet = wb.active

        cities = self.settings["Аббревиатуры городов"]

        # Вставляем данные в новый файл, начиная с определенной строки
        row_n = 1  # Начинаем вставку со 2-й строки

        # Получаем заголовки из первой строки, для определения колонок
        headers = [cell.value for cell in sheet[1]]

        marka_value = ""
        model_value = ""
        pg_value = ""

        # Создаем словарь кодов городов, где ключ — Код, значение — Аббревиатура
        city_params_dict = {}
        city_code_dict = {}
        city_api_dict = {}
        for city, details in cities.items():
            code = details["Код"]
            abbreviation = details["Аббревиатура"]
            api_key = details["Обновление по API"]
            city_code_dict[code] = abbreviation
            city_api_dict[abbreviation] = api_key
        city_params_dict['codes'] = city_code_dict
        city_params_dict['api_keys'] = city_api_dict

        # Преобразование словаря в строку JSON
        json_str = json.dumps(city_params_dict)

        # Настройки столбцов
        for col_index, header in enumerate(headers):
            if header == "api_json":
                sheet.cell(row=1, column=col_index + 1).value = json_str
            # Скрываем столбцы
            if header in hidden_cols:
                sheet.column_dimensions[sheet.cell(row=1, column=col_index + 1).column_letter].hidden = True
        

        # Вставляем данные
        for item in price:
            # if item["Артикул"] not in [None, '', 0]:
            row_n += 1  # Переходим к следующей строке
            # Определяем формат границ
            if item["Марка"] != marka_value:
                new_border = marka_border
                marka_value = item["Марка"]
                model_value = item["Модель"]
                pg_value = item["Товарная группа"]
            elif item["Модель"] != model_value:
                new_border = model_border
                model_value = item["Модель"]
                pg_value = item["Товарная группа"]
            elif (item["Товарная группа"] != pg_value 
                  and not (item["Товарная группа"].startswith("Брызговики") 
                           and pg_value.startswith("Брызговики"))):
                new_border = pg_border
                model_value = item["Модель"]
                pg_value = item["Товарная группа"]
            else:
                new_border = thin_border
            for col_n in range(1, sheet.max_column + 1):
                sheet.cell(row=row_n, column=col_n).border = new_border
            # Заливка ячеек
            for col_index, header in enumerate(headers):
                col_n = col_index + 1
                if header == "Прайс":
                    sheet.cell(row=row_n, column=col_n).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2',
                                                        fill_type='solid')
            # Вставка данных
            for col_index, header in enumerate(headers):
                col_n = col_index + 1
                # Проверяем, есть ли заголовок в данных и не пустые ли они
                if (header in item) and (item[header] not in [None, '', 0]):  
                    # Записываем данные в соответствующую ячейку
                    sheet.cell(row=row_n, column=col_n).value = item[header]
                    if header == "Прайс": 
                        if item["AvitoStatus"] == "Активно":
                            sheet.cell(row=row_n, column=col_n).fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE',
                                                                            fill_type='solid')
                        else:
                            sheet.cell(row=row_n, column=col_n).fill = PatternFill(start_color='F2DCD7', end_color='F2DCD7',
                                                                            fill_type='solid')
                # Добавляем форматирование для городов
                for city in cities:
                    if header == cities[city]["Аббревиатура"]:
                        new_color = cities[city]["Цвет"]
                        sheet.cell(row=row_n, column=col_n).fill = PatternFill(start_color=new_color,
                                                                                end_color=new_color,
                                                                                fill_type='solid') 
                        sheet.cell(row=row_n, column=col_n).alignment = Alignment(horizontal='center')
                # Границы ячейки
                sheet.cell(row=row_n, column=col_n).border = new_border               

        # Сохраняем изменения в новом файле
        wb.save(price_file_path)

        print(f"Файл 'Прайс.xlsm' успешно сформирован.")


       
    
    def copy_ostatki_to_price(self):
        # Копируем шаблон в новую директорию
        price_template = self.settings["Настройки"]["Шаблон_для_Прайса"]
        ostatki_path = self.settings["Настройки"]["Остатки"]
        price_path = self.settings["Настройки"]["Прайс_для_работы"]

        shutil.copy(price_template, price_path)
        print(f"Данные из {price_template} успешно скопированы в {
              price_path}.")

        # Загрузка книги с остатками
        ostatki_workbook = load_workbook(ostatki_path)
        ostatki_sheet = ostatki_workbook.active  # Выбираем активный лист

        # Загрузка книги прайса
        price_workbook = load_workbook(
            filename=price_path, read_only=False, keep_vba=True)
        price_sheet = price_workbook.active  # Выбираем активный лист

        # Получение первой строки
        first_row = next(ostatki_sheet.iter_rows(values_only=True))

        # Вставка значений в первую строку price_sheet
        for col_index, value in enumerate(first_row):
            price_sheet.cell(row=1, column=col_index + 1, value=value)

        # # Итерация по строкам
        for index, row in enumerate(ostatki_sheet.iter_rows(values_only=True)):
            if index > 1:  # Пропустить вторую строку (индекс 1)
                # Определяем строку для вставки
                target_row = index - 1  # Вставляем с первой строки, смещая на одну вверх
                for col_index, value in enumerate(row):
                    # Вставляем значение в соответствующую ячейку
                    price_sheet.cell(row=target_row + 1,
                                     column=col_index + 1, value=value)

        # Сохранение изменений в price_path
        price_workbook.save(price_path)

        print(f"Данные из {ostatki_path} успешно скопированы в {price_path}.")

    def price_format(self):
        price_path = self.settings["Настройки"]["Прайс_для_работы"]

        # Создаем новую книгу, если прайс еще не существует
        try:
            workbook = load_workbook(
                filename=price_path, read_only=False, keep_vba=True)
        except FileNotFoundError:
            workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Прайс"

        # # Сначала создаем список объединенных диапазонов
        # merged_ranges_to_unmerge = []
        # for merged_range in sheet.merged_cells.ranges:
        #     if merged_range.bounds[1] <= 2 <= merged_range.bounds[3]:  # Проверяем, включает ли объединение 2-ю строку
        #         merged_ranges_to_unmerge.append(merged_range)
        #
        # # Теперь разъединяем ячейки
        # for merged_range in merged_ranges_to_unmerge:
        #     sheet.unmerge_cells(str(merged_range))  # Разъединяем ячейки
        #
        # # Теперь можем безопасно удалить 2-ю строку
        # sheet.delete_rows(2)

        # Меняем местами столбцы 7 и 8 и форматируем числа
        for row in range(1, sheet.max_row + 1):
            col_value = sheet.cell(row=row, column=7).value
            col_number_format = sheet.cell(row=row, column=7).number_format
            # Меняем местами значения
            sheet.cell(row=row, column=7).value = sheet.cell(
                row=row, column=8).value
            sheet.cell(row=row, column=8).value = col_value
            # Устанавливаем числовой формат с разделителем и без нулей после десятичной точки
            sheet.cell(row=row, column=7).number_format = sheet.cell(
                row=row, column=8).number_format
            sheet.cell(row=row, column=8).number_format = col_number_format
            # sheet.cell(row=row, column=7).number_format = '#,##0.##'
            # sheet.cell(row=row, column=8).number_format = '#,##0.##'
        # Переименовываем столбцы 7 и 8
        sheet.cell(row=1, column=7).value = "Цена"
        sheet.cell(row=1, column=8).value = "Расп"

        # Получаем настройки аббревиатур городов
        abbreviations = self.settings["Аббревиатуры городов"]

        # Проходим по всем ячейкам в первой строке
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            # Проверяем, есть ли значение в настройках
            if cell_value in abbreviations:
                # Заменяем значение на аббревиатуру
                sheet.cell(
                    row=1, column=col).value = abbreviations[cell_value]["Аббревиатура"]
                sheet.column_dimensions[sheet.cell(
                    row=1, column=col).column_letter].width = 7  # Уменьшаем размер
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')

        # Сортируем города по их "Коду"
        sorted_cities = sorted(abbreviations.items(),
                               key=lambda x: int(x[1]["Сортировка"]))

        # Определяем, на каком столбце будем добавлять новые
        start_col = sheet.max_column + 1
        cols_to_delete = []
        # Добавляем столбцы для каждого города
        for index, (city, data) in enumerate(sorted_cities):
            column_letter = start_col + index
            cell = sheet.cell(row=1, column=column_letter,
                              value="Аббревиатура")
            cell.value = data["Аббревиатура"]
            new_color = data["Цвет"]

            # Проходим по первой строке до start_col и ищем совпадения
            for col in range(1, start_col):
                if sheet.cell(row=1, column=col).value == data["Аббревиатура"]:
                    cols_to_delete.append(col)
                    # Копируем значения из найденного столбца
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row=row, column=column_letter).value = sheet.cell(
                            row=row, column=col).value
                        sheet.cell(row=row, column=column_letter).fill = PatternFill(start_color=new_color,
                                                                                     end_color=new_color,
                                                                                     fill_type='solid')
                        sheet.cell(row=row, column=column_letter).alignment = Alignment(horizontal='center')
        # Опять проходим по первой строке до start_col и удаляем старые столбцы городов
        cols_to_delete.sort(reverse=True)
        # print(cols_to_delete)
        for col in cols_to_delete:
            sheet.delete_cols(col, 1)  # Удаляем столбец

        # Получаем номер последнего столбца
        last_col = sheet.max_column

        # Копируем столбцы 3-6 в конец таблицы
        for col in range(3, 7):  # Столбцы 3, 4, 5, 6
            for row in range(1, sheet.max_row + 1):
                sheet.cell(row=row, column=last_col + (col - 3) +
                           1).value = sheet.cell(row=row, column=col).value

        # Удаляем исходные столбцы 3-6
        sheet.delete_cols(3, 4)  # Удаляем 4 столбца, начиная с 3-го

        # Добавляем столбцы: 1 - "Вес", 2 и 3 - пустые, 4 - "ID_01"
        sheet.insert_cols(1, 5)  # Вставляем 4 столбца (1, 2, 3, 4)
        sheet.cell(row=1, column=1, value="Вес")
        sheet.cell(row=1, column=2, value="")
        sheet.cell(row=1, column=3, value="")
        sheet.cell(row=1, column=4, value="")
        sheet.cell(row=1, column=5, value="ID_01")

        # Устанавливаем ширину столбцов
        # sheet.column_dimensions[sheet.cell(
        #     row=1, column=1).column_letter].width = 4.17  # 1-й столбец (Вес)
        # sheet.column_dimensions[sheet.cell(
        #     row=1, column=2).column_letter].width = 0.82  # 2-й столбец
        # sheet.column_dimensions[sheet.cell(
        #     row=1, column=3).column_letter].width = 0.82  # 3-й столбец
        # sheet.column_dimensions[sheet.cell(
        #     row=1, column=4).column_letter].width = pixels_to_width(9)  # 4-й столбец
        sheet.column_dimensions[sheet.cell(
            row=1, column=5).column_letter].hidden = True  # 5-й столбец (ID_01)

        # Добавляем столбцы с параметрами из раздела "Авито"
        avito_settings = self.settings["Авито"]  # Обращаемся к разделу "Авито"
        # Начинаем с 6-го столбца (5 - потому что будем сразу увеличивать на 1)
        col_n = 5

        for key, value in avito_settings.items():
            col_n += 1  # Увеличиваем смещение для следующего столбца
            column_name = f"Avitoid_{value['Код']}"
            # Вставляем новый столбец перед добавлением заголовка
            sheet.insert_cols(col_n)
            # Заполняем заголовок
            sheet.cell(row=1, column=col_n, value=column_name)

        # Устанавливаем ширину для дополнительных столбцов
        for i in range(5, col_n + 1):  # 5-й столбец и дальше
            # Ширина по вашему выбору
            # sheet.column_dimensions[sheet.cell(
            #     row=1, column=i).column_letter].width = 11
            sheet.column_dimensions[sheet.cell(
                row=1, column=i).column_letter].hidden = True

        col_n += 1
        sheet.column_dimensions[sheet.cell(
            row=1, column=col_n).column_letter].width = pixels_to_width(145)  # Размер для Артикул
        col_n += 1
        sheet.column_dimensions[sheet.cell(
            row=1, column=col_n).column_letter].width = pixels_to_width(790)  # Размер для Товар
        col_n += 1

        # Создаем словарь кодов городов, где ключ — Код, значение — Аббревиатура
        city_params_dict = {}
        city_code_dict = {}
        city_api_dict = {}
        for city, details in abbreviations.items():
            code = details["Код"]
            abbreviation = details["Аббревиатура"]
            api_key = details["Обновление по API"]
            city_code_dict[code] = abbreviation
            city_api_dict[abbreviation] = api_key
        city_params_dict['codes'] = city_code_dict
        city_params_dict['api_keys'] = city_api_dict

        # Преобразование словаря в строку JSON
        json_str = json.dumps(city_params_dict)

        sheet.insert_cols(col_n)  # Добавляем пустой столбец
        # sheet.cell(row=1, column=col_n, value="")
        # В заголовок записываем справочник кодов городов
        sheet.cell(row=1, column=col_n).value = json_str
        sheet.column_dimensions[sheet.cell(
            row=1, column=col_n).column_letter].width = 2

        col_n += 1
        # Вставляем новый столбец перед добавлением заголовка
        sheet.insert_cols(col_n)
        sheet.cell(row=1, column=col_n, value="Прайс")  # Заполняем заголовок
        # Заливаем столбец Прайс цветом (242,242,242)
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=col_n).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2',
                                                                 fill_type='solid')
        sheet.column_dimensions[sheet.cell(
            row=1, column=col_n).column_letter].width = 10  # Размер для Прайс
        col_n += 1
        sheet.column_dimensions[sheet.cell(
            row=1, column=col_n).column_letter].width = 10  # Размер для Цена
        
        col_n += 1
        # Вставляем новый столбец -10% перед добавлением заголовка
        sheet.insert_cols(col_n)
        sheet.cell(row=1, column=col_n, value="-10%")  # Заполняем заголовок
        for row in range(2, sheet.max_row + 1):
            price_str = sheet.cell(row=row, column=col_n-1).value
            price_val = 0
            try:
                if price_str is not None:  # Проверяем, что ячейка не пуста
                    price_val = float(price_str)
            except ValueError:
                pass
            price_10_val = round(price_val * 0.9, 2)
            sheet.cell(row=row, column=col_n).value = price_10_val
        sheet.column_dimensions[get_column_letter(col_n)].hidden = True  # Скрываем столбец -10%

        col_n += 1
        sheet.column_dimensions[sheet.cell(
            row=1, column=col_n).column_letter].width = 10  # Размер для Расп
        # col_n += 1
        # sheet.insert_cols(col_n)  # Вставляем новый столбец перед добавлением заголовка

        # # Создаем словарь кодов городов, где ключ — Код, значение — Аббревиатура
        # city_params_dict = {}
        # city_code_dict = {}
        # city_api_dict = {}
        # for city, details in abbreviations.items():
        #     code = details["Код"]
        #     abbreviation = details["Аббревиатура"]
        #     api_key = details["Обновление по API"]
        #     city_code_dict[code] = abbreviation
        #     city_api_dict[abbreviation] = api_key
        # city_params_dict['codes'] = city_code_dict
        # city_params_dict['api_keys'] = city_api_dict

        # # Преобразование словаря в строку JSON
        # json_str = json.dumps(city_params_dict)

        # sheet.cell(row=1, column=col_n).value = json_str  # В заголовок записываем справочник кодов городов
        # sheet.column_dimensions[sheet.cell(row=1, column=col_n).column_letter].width = 2

        # Устанавливаем размер шрифта в первой строке
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col).font = Font(size=10)

        # Заливаем первую строку цветом (79,129,189) и устанавливаем белый цвет шрифта
        header_fill = PatternFill(
            start_color='4F7DBE', end_color='4F7DBE', fill_type='solid')
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col).fill = header_fill
            sheet.cell(row=1, column=col).font = Font(
                color='FFFFFF')  # Белый цвет

        # Устанавливаем сетку с цветом (79,129,189)
        border_color = Side(style='thin', color='4F7DBE')
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(
                    left=border_color, right=border_color, top=border_color, bottom=border_color)

        # Замораживаем первую строку
        sheet.freeze_panes = sheet["A2"]

        # Сохраняем изменения
        workbook.save(price_path)
        print(f"Таблица прайса успешно сформирована и сохранена в '{
              price_path}'.")

    def update_price_from_goods(self):
        price_path = self.settings["Настройки"]["Прайс_для_работы"]
        goods_path = self.settings["Настройки"]["Таблица_с_товарами"]

        link_0 = self.settings["Ссылки"]["Ссылка_прайса"]
        # link_1 = self.settings["Ссылки"]["Ссылка_прайса_1"]
        # link_2 = self.settings["Ссылки"]["Ссылка_прайса_2"]

        # Загружаем таблицы
        price_workbook = load_workbook(
            filename=price_path, read_only=False, keep_vba=True)
        goods_workbook = load_workbook(goods_path)

        price_sheet = price_workbook.active
        goods_sheet = goods_workbook.active

        # Находим индекс столбца "Артикул" в таблице "Прайс"
        article_col_price = None
        for col in range(1, price_sheet.max_column + 1):
            if price_sheet.cell(row=1, column=col).value == "Артикул":
                article_col_price = col
                break

        # Находим индекс столбца "Бренд" в таблице "Прайс"
        brand_col_price = None
        for col in range(1, price_sheet.max_column + 1):
            if price_sheet.cell(row=1, column=col).value == "Бренд":
                brand_col_price = col
                break

         # Находим индекс столбца "Товарная группа" в таблице "Прайс"
        product_group_col_price = None
        for col in range(1, price_sheet.max_column + 1):
            if price_sheet.cell(row=1, column=col).value == "Товарная группа":
                product_group_col_price = col
                break

        # Находим индекс столбца "Артикул" в таблице "Товар"
        article_col_goods = None
        for col in range(1, goods_sheet.max_column + 1):
            if goods_sheet.cell(row=1, column=col).value == "Артикул":
                article_col_goods = col
                break

        if article_col_price is None or article_col_goods is None:
            print("Не удалось найти столбец 'Артикул' в одной из таблиц.")
            return

        avito_settings = self.settings["Авито"]  # Обращаемся к разделу "Авито"
        avito_count = len(avito_settings)

        # Создаем словарь для значений из таблицы "Товар"
        goods_data = self.excel_to_dict(
            excel_path=self.settings["Настройки"]["Таблица_с_товарами"], key_col_header="Артикул", key_upper=True)
        for key in goods_data:
            goods_data[key]["is_correct"] = False
            if goods_data[key]["Артикул"] is not None:
                goods_data[key]["Артикул"] = goods_data[key]["Артикул"].upper()
                for code in range(1, avito_count + 1):
                    if goods_data[key][f"{code:02}: AvitoId"] != "":
                        goods_data[key]["is_correct"] = True

        # Скачиваем XML-файл
        print('Начало скачивания файла XML')
        xml_path = self.settings["Настройки"]["XML_файл"]
        download_xml(
            'http://portal.autofamily.ru/export/autofamily_catalog.xml', xml_path)
        print('Скачивание файла XML завершено')

        # Парсим XML-файл
        article_weight = parse_xml(xml_path)

        # Вставляем веса и значения из таблицы "Товар" в таблицу "Прайс"

        start_col = 5  # Начинаем с 5-го столбца (номер 5)
        for row in range(2, price_sheet.max_row + 1):
            article_value = str(price_sheet.cell(
                row=row, column=article_col_price).value)
            article_value = article_value.upper()
            # Вставляем веса
            if article_value in article_weight:
                weight = article_weight[article_value]
                # Записываем вес в 1-й столбец
                price_sheet.cell(row=row, column=1, value=weight)
            # Вставляем значения из таблицы "Товар"
            if (article_value in goods_data) and goods_data[article_value]["is_correct"]:
                col_n = start_col
                price_sheet.cell(
                    row=row, column=col_n).value = goods_data[article_value]["01: Id"]
                for code in range(1, avito_count + 1):
                    col_n += 1
                    price_sheet.cell(row=row, column=col_n).value = goods_data[article_value][f"{
                        code:02}: AvitoId"]
                col_n += 4
                price_sheet.cell(
                    row=row, column=col_n).value = goods_data[article_value]["01: Price"]
                if goods_data[article_value]["AvitoStatus"] == "Активно":
                    price_sheet.cell(row=row, column=col_n).fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE',
                                                                               fill_type='solid')
                else:
                    price_sheet.cell(row=row, column=col_n).fill = PatternFill(start_color='F2DCD7', end_color='F2DCD7',
                                                                               fill_type='solid')
                # Вставляем ссылки
                # price_sheet.cell(row=row, column=2).value = link_1.replace("{ID_01}",
                #                                                            goods_data[article_value]["ID_01"])
                # price_sheet.cell(row=row, column=3).value = link_2.replace("{ID_01}",
                #                                                            goods_data[article_value]["ID_01"])

                for code in range(1, avito_count + 1):
                    price_sheet.cell(row=row, column=1+code).value = link_0.replace(
                        "{ID_XX}", str(goods_data[article_value][f"{code:02}: AvitoId"]))

        self.apply_filters_to_price_sheet(price_sheet, article_col_price, brand_col_price, product_group_col_price)

        # Сохраняем изменения в таблице "Прайс"
        price_workbook.save(price_path)
        print(f"Таблица 'Прайс' обновлена на основе таблицы 'Товар'.")

    def apply_filters_to_price_sheet(self, sheet, article_col, brand_col, product_group_col):
        # Получаем фильтры из настроек
        text_filters = self.settings["Фильтры"]["Фильтр_по_тексту"]
        article_filters = self.settings["Фильтры"]["Фильтр_по_артикулу"]
        group_brand_filters = self.settings["Фильтры"]["Фильтр_по_товарная_группа_и_бренд"]

        # Удаляем вхождения текстовых фильтров из 9 столбца
        for row in range(2, sheet.max_row + 1):  # Начинаем с 2-й строки
            cell_value = sheet.cell(row=row, column=article_col + 1).value
            if cell_value:  # Проверяем, что значение не None
                for text_filter in text_filters:
                    cell_value = cell_value.replace(
                        text_filter, "")  # Заменяем на пустую строку
                # Убираем лишние пробелы
                sheet.cell(row=row, column=article_col +
                           1).value = cell_value.strip()

        # Помечаем на удаление строки, если найдены соответствия по article_col
        rows_to_delete = []
        for row in range(2, sheet.max_row + 1):
            article_value = sheet.cell(row=row, column=article_col).value
            brand_value = sheet.cell(row=row, column=brand_col).value
            product_group_value = sheet.cell(row=row, column=product_group_col).value
            if (article_value in article_filters) or (brand_value in group_brand_filters) or (product_group_value in group_brand_filters):
                rows_to_delete.append(row)

        # Помечаем на удаление строки, если найдены соответствия по brand_col или product_group_col
        # rows_to_delete = []
        # for row in range(2, sheet.max_row + 1):
        #     brand_value = sheet.cell(row=row, column=brand_col).value
        #     product_group_value = sheet.cell(row=row, column=product_group_col).value
        #     if (brand_value in group_brand_filters) or (product_group_value in group_brand_filters):
        #         rows_to_delete.append(row)

        # Удаляем строки, начиная с конца, чтобы не нарушить индексацию
        for row in reversed(rows_to_delete):
            sheet.delete_rows(row)

    def create_price_file(self, check_email=True):
        if check_email:
            self.load_ostatki()
        self.create_price_xlsx()
        # self.copy_ostatki_to_price()
        # self.price_format()
        # self.update_price_from_goods()

    def excel_to_list(self, excel_path):
        # Открываем Excel-файл
        workbook = load_workbook(excel_path)
        sheet = workbook.active  # Получаем активный лист

        # Get the headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Remove columns with empty headers
        valid_columns = [i for i, header in enumerate(headers) if header]

        # Create an empty list to store the result
        result_list = []

        # Iterate over each row starting from the second row (data rows)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in valid_columns}
            result_list.append(row_dict)
        return result_list

    def excel_to_dict(self, excel_path, key_col_header, key_upper=False):
        # Открываем Excel-файл
        workbook = load_workbook(excel_path)
        sheet = workbook.active  # Получаем активный лист

        # Get the headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Remove columns with empty headers
        valid_columns = [i for i, header in enumerate(headers) if header]

        # Create an empty dictionary to store the result
        result_dict = {}

        # Iterate over each row starting from the second row (data rows)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the "Артикул" value for this row to use as the key
            art_key = row[headers.index(
                key_col_header)] if key_col_header in headers else None

            if art_key:
                # Create a dictionary for this row, including only valid columns
                # row_dict = {headers[i]: row[i] for i in valid_columns if headers[i] != key_col_header}
                if key_upper and isinstance(art_key, str):
                    art_key = art_key.upper()
                row_dict = {headers[i]: row[i] for i in valid_columns}
                # Add the row dictionary to the main result dictionary
                result_dict[art_key] = row_dict

        return result_dict

    def create_orders_files(self, stores=[], check_email=False):
        # Проверяем нужно ли загружать остатки по email
        if check_email:
            self.load_ostatki()
        # Отбираем потребность (данные по товарам)
        # goods_data = self.excel_to_list(excel_path=self.settings["Настройки"]["Таблица_с_товарами"])
        goods_data = self.excel_to_dict(
            excel_path=self.settings["Настройки"]["Таблица_с_товарами"], key_col_header="Артикул", key_upper=True)
        # keys_to_keep = ["Артикул", "02: К", "03: Р"]
        orders_data = {}
        for key in goods_data:
            n2 = 0
            n3 = 0
            if stores and ("02" in stores):
                n2 = read_int(goods_data[key]['02: К'])
            if stores and ("03" in stores):
                n3 = read_int(goods_data[key]['03: Р'])
            if (n2 > 0) or (n3 > 0):
                article_need = {"Артикул": key, '02_need': n2, '03_need': n3}
                orders_data[key] = article_need

        # for item in goods_data:
        #     n2 = read_int(item['02: К'])
        #     n3 = read_int(item['03: Р'])
        #     if (n2 > 0) or (n3 > 0):
        #         article_need = {"Артикул": item["Артикул"], '02_need': n2, '03_need': n3}
        #         orders_data[item["Артикул"]] = article_need
        # pprint(orders_data)

        # Считываем перемещения
        for id in stores:
            source_file = self.settings["Авито"][f"{id}"]["Файл перемещения"]
            source_data = self.excel_to_dict(
                excel_path=source_file, key_col_header="Артикул", key_upper=True)
            for item in orders_data:
                if item in source_data:
                    # print(item)
                    orders_data[item][f'{id}_mov'] = read_int(
                        source_data[item]['Количество'])
                else:
                    orders_data[item][f'{id}_mov'] = 0

        # Считываем наличие
        if check_email:
            # source_file = self.settings["Настройки"]["Прайс_для_работы"]
            source_file = self.settings["Настройки"]["Остатки"]
            source_data = self.excel_to_dict(
                excel_path=source_file, key_col_header="Артикул", key_upper=True)
            for item in orders_data:
                if item in source_data:
                    orders_data[item]["01_avail"] = read_int(
                        source_data[item]["Распределительный Центр 3"])
                    orders_data[item]["02_avail"] = read_int(
                        source_data[item]["Краснодар"])
                    orders_data[item]["03_avail"] = read_int(
                        source_data[item]["Ростов-на-Дону"])
                    # orders_data[item]["01_avail"] = read_int(source_data[item]["МСК"])
                    # orders_data[item]["02_avail"] = read_int(source_data[item]["КРД"])
                    # orders_data[item]["03_avail"] = read_int(source_data[item]["РСТ"])
                    orders_data[item]["Товар"] = source_data[item]["Товар"]
                else:
                    orders_data[item]["01_avail"] = 0
                    orders_data[item]["02_avail"] = 0
                    orders_data[item]["03_avail"] = 0
                    orders_data[item]["Товар"] = ""
        else:
            # Считываем наличие по API
            for item in orders_data:
                warehauses_data = get_warehouses_data(
                    article=orders_data[item]["Артикул"], api_key=AUTOFAMILY_API_KEY)
                if warehauses_data:
                    orders_data[item]["01_avail"] = warehauses_data["warehouses"][self.settings["Аббревиатуры городов"]
                                                                                ["Распределительный Центр 3"]["Код"]]
                    orders_data[item]["02_avail"] = warehauses_data["warehouses"][self.settings["Аббревиатуры городов"]["Краснодар"]["Код"]]
                    orders_data[item]["03_avail"] = warehauses_data["warehouses"][self.settings["Аббревиатуры городов"]
                                                                                ["Ростов-на-Дону"]["Код"]]
                    orders_data[item]["Товар"] = warehauses_data["name"]
                else:
                    orders_data[item]["01_avail"] = 0
                    orders_data[item]["02_avail"] = 0
                    orders_data[item]["03_avail"] = 0
                    orders_data[item]["Товар"] = ""

        # Обрабатываем данные
        for item in orders_data:
            av_01 = orders_data[item]["01_avail"]
            lack_02 = 0
            lack_03 = 0
            if "02" in stores:
                lack_02 = orders_data[item]["02_need"] - \
                    orders_data[item]["02_avail"] - orders_data[item]["02_mov"]
            if "03" in stores:
                lack_03 = orders_data[item]["03_need"] - \
                    orders_data[item]["03_avail"] - orders_data[item]["03_mov"]
            def_02 = lack_02
            def_03 = lack_03

            order_02 = 0
            order_03 = 0

            while (av_01 > 0) and ((def_03 > 0) or (def_02 > 0)):
                if def_03 > 0 and av_01 > 0:
                    order_03 += 1
                    def_03 -= 1
                    av_01 -= 1
                if def_02 > 0 and av_01 > 0:
                    order_02 += 1
                    def_02 -= 1
                    av_01 -= 1

            orders_data[item]["02_lack"] = lack_02
            orders_data[item]["03_lack"] = lack_03
            orders_data[item]["02_order"] = order_02
            orders_data[item]["03_order"] = order_03
            orders_data[item]["02_def"] = def_02
            orders_data[item]["03_def"] = def_03

        # pprint(orders_data)

        # Формируем Excel файлы
        orders_template = self.settings["Настройки"]["Шаблон_для_Заказов"]
        # Получаем фильтры из настроек
        text_filters = self.settings["Фильтры"]["Фильтр_по_тексту"]
        for id in stores:
            new_file_path = self.settings["Авито"][f"{id}"]["Файл заказов"]
            shutil.copy(orders_template, new_file_path)

            # Открываем скопированный файл
            wb = load_workbook(new_file_path)
            sheet = wb.active

            # Вставляем данные в новый файл, начиная с определенной строки
            row_n = 1  # Начинаем вставку со 2-й строки

            # Получаем заголовки из первой строки, чтобы сопоставить с orders_data
            headers = [cell.value for cell in sheet[1]]
            # orders_temp = [val for key, val in orders_data.items() if val[f"{id}_def"] == 0]
            # orders_groups = [orders_temp]
            # orders_temp = [val for key, val in orders_data.items() if (val[f"{id}_def"] > 0) and (val["01_avail"] > 0)] 
            # orders_groups.append(orders_temp) 
            # orders_temp = [val for key, val in orders_data.items() if (val[f"{id}_def"] > 0) and (val["01_avail"] == 0)] 
            # orders_groups.append(orders_temp) 

            orders = []
            for key, val in orders_data.items():
                if val[f"{id}_def"] == 0:
                    orders.append((val, ""))
                elif (val[f"{id}_def"] > 0) and (val["01_avail"] > 0):
                    orders.append((val, COLORS["YELLOW"]))
            orders_temp = [(val, COLORS["RED"]) for key, val in orders_data.items() if (val[f"{id}_def"] > 0) and (val["01_avail"] == 0)] 
            orders += orders_temp  

            for item in orders:
                order = item[0]
                cell_color = item[1]
                if order[f"{id}_lack"] > 0:
                    # Добавляем строку в Excel
                    # orders_data[item][f'{id}_mov'] = read_int(source_data[item]['Количество'])
                    nomenclature = order["Товар"]
                    for text_filter in text_filters:
                        nomenclature = nomenclature.replace(
                            text_filter, "")  # Заменяем на пустую строку
                    nomenclature = nomenclature.strip()  # Убираем лишние пробелы

                    row_n += 1  # Переходим к следующей строке
                    for col_index, header in enumerate(headers):
                        col_n = col_index + 1
                        if header == "Артикул":
                            sheet.cell(row=row_n, column=col_n).value = order["Артикул"]
                        elif header == "Требуется":
                            sheet.cell(row=row_n, column=col_n).value = order[f"{id}_need"]
                        elif header == "Наличие":
                            sheet.cell(row=row_n, column=col_n).value = order[f"{id}_avail"]
                        elif header == "Перемещ.":
                            sheet.cell(row=row_n, column=col_n).value = order[f"{id}_mov"]
                        elif header == "Москва, всего":
                            sheet.cell(row=row_n, column=col_n).value = order["01_avail"]
                            if cell_color:
                                sheet.cell(row=row_n, column=col_n).fill = PatternFill(start_color=cell_color, end_color=cell_color,
                                                                                    fill_type='solid')
                            else:
                                sheet.cell(row=row_n, column=col_n).fill = PatternFill()
                        elif header == "Москва, дефицит":
                            sheet.cell(row=row_n, column=col_n).value = order[f"{id}_def"]
                            if cell_color:
                                sheet.cell(row=row_n, column=col_n).fill = PatternFill(start_color=cell_color, end_color=cell_color,
                                                                                    fill_type='solid')
                            else:
                                sheet.cell(row=row_n, column=col_n).fill = PatternFill()
                        elif header == "Заказать":
                            sheet.cell(row=row_n, column=col_n).value = order[f"{id}_lack"]
                        elif header == "Номенклатура":
                            sheet.cell(row=row_n, column=col_n).value = nomenclature

            # Сохраняем изменения в новом файле
            wb.save(new_file_path)

    def calculate_article_sums(self):
        price_path = self.settings["Настройки"]["Прайс_для_работы"]

        # Загружаем таблицы
        price_workbook = load_workbook(
            filename=price_path, read_only=False, keep_vba=True)
        sheet = price_workbook.active  # Получаем активный лист

        # Получаем заголовки из первой строки
        headers = [cell.value for cell in sheet[1]]

        # Находим индексы необходимых столбцов
        article_index = headers.index("Артикул")
        brand_index = headers.index("Марка")

        result = {}

        # Итерируемся по строкам, начиная со второй
        for row in sheet.iter_rows(min_row=2, values_only=True):
            article = row[article_index]  # Получаем артикул

            # Суммируем значения от 6 ячеек правее до столбца "Марка", игнорируя None
            total_sum = sum(
                value for value in row[article_index + 6:brand_index] if isinstance(value, (int, float)))

            result[article] = total_sum  # Записываем в словарь

        return result

    def create_avito_file(self, code, goods_data, article_sums, only_active=True):
        # Копируем шаблон в новую директорию
        avito_template = self.settings["Настройки"]["Шаблон_для_Авито"]
        if only_active:
            new_file_path = self.settings["Авито"][code]["Путь_сохранения_Только_активные"]
        else:
            new_file_path = self.settings["Авито"][code]["Путь_сохранения_Все_товары"]
        shutil.copy(avito_template, new_file_path)

        # Открываем скопированный файл
        wb = load_workbook(new_file_path)
        sheet = wb.active

        # Вставляем данные в новый файл, начиная с определенной строки
        row_n = 1  # Начинаем вставку со 2-й строки

        # Получаем заголовки из первой строки, чтобы сопоставить с row_data
        headers = [cell.value for cell in sheet[1]]

        # Считываем из настроек параметры
        email = self.settings["Авито"][code]["EMail"]
        company_name = self.settings["Авито"][code]["CompanyName"]
        category = self.settings["Авито"][code]["Category"]

        avito_date_end_str = (date.today() + timedelta(days=30)).strftime("%Y-%m-%dT00:00:00")

        for row_data in goods_data:
            if row_data[f'{code}: AvitoId'] not in [None, '', 0]:
                row_n += 1  # Переходим к следующей строке
                for col_index, header in enumerate(headers):
                    col_n = col_index + 1
                    if header in row_data:  # Проверяем, есть ли заголовок в row_data
                        # Записываем данные в соответствующую ячейку
                        sheet.cell(
                            row=row_n, column=col_n).value = row_data[header]
                    # По коду ищем нужные данные
                    elif f'{code}: {header}' in row_data:
                        # Записываем данные в соответствующую ячейку
                        sheet.cell(row=row_n, column=col_n).value = row_data[f'{
                            code}: {header}']
                    elif header == 'EMail':
                        sheet.cell(row=row_n, column=col_n).value = email
                    elif header == 'CompanyName':
                        sheet.cell(
                            row=row_n, column=col_n).value = company_name
                    elif header == 'Category':
                        sheet.cell(row=row_n, column=col_n).value = category
                    elif header == 'ListingFee':
                        sheet.cell(row=row_n, column=col_n).value = "Package"
                    elif header == 'Condition':
                        sheet.cell(row=row_n, column=col_n).value = "Новое"
                    elif header == 'AdType':
                        sheet.cell(row=row_n, column=col_n).value = "Товар от производителя"
                    elif header == 'AvitoDateEnd':
                        sheet.cell(row=row_n, column=col_n).value = avito_date_end_str
                # Проверяем AvitoStatus
                if sheet.cell(row=row_n, column=4).value == "Активно":
                    if sheet.cell(row=row_n, column=1).value in article_sums:
                        if article_sums[sheet.cell(row=row_n, column=1).value] == 0:
                            sheet.cell(
                                row=row_n, column=4).value = "Снято с публикации"
                    else:
                        sheet.cell(
                            row=row_n, column=4).value == "Снято с публикации"

        # Удаляем то, что снято с публикации
        if only_active:
            rows_to_delete = []

            # Ищем строки для удаления
            for row in sheet.iter_rows(min_row=2):  # Начинаем со второй строки
                # Проверяем 4-й столбец (индекс 3)
                if row[3].value != "Активно":
                    # Сохраняем номер строки для удаления
                    rows_to_delete.append(row[0].row)

            # Удаляем строки в обратном порядке
            for row_index in reversed(rows_to_delete):
                sheet.delete_rows(row_index)

        # Удаляем столбец артикула
        sheet.delete_cols(1, 1)  # Удаляем столбец

        # Сохраняем изменения в новом файле
        wb.save(new_file_path)

    def create_all_avito_files(self, only_active=True):
        goods_data = self.excel_to_list(
            excel_path=self.settings["Настройки"]["Таблица_с_товарами"])
        article_sums = self.calculate_article_sums()
        # Получаем список ключей из раздела "Авито"
        avito_keys = list(self.settings["Авито"].keys())
        for code in avito_keys:
            self.create_avito_file(code=code, goods_data=goods_data,
                                   article_sums=article_sums, only_active=only_active)

    def create_price_weigh_file(self):
        # Параметры
        price_path = self.settings["Настройки"]["Прайс_для_работы"]
        price_and_weight_template = self.settings["Настройки"]["Шаблон_для_Цены_и_вес"]
        out_dir = self.settings["Настройки"]["Папка_с_выходными_данными"]
        new_file_name = "Цены и вес.xlsx"
        new_file_path = os.path.join(out_dir, new_file_name)

        # Открываем файл с прайсом
        wb_price = load_workbook(
            filename=price_path, read_only=False, keep_vba=True)
        sheet_price = wb_price.active

        # Находим индексы необходимых столбцов
        headers = [cell.value for cell in sheet_price[1]]
        id_index = headers.index("ID_01")
        article_index = headers.index("Артикул")
        price_index = headers.index("Цена")
        weight_index = headers.index("Вес")
        product_index = headers.index("Товар")

        # Копируем значения
        data = []
        for row in sheet_price.iter_rows(min_row=2, values_only=True):
            if row[id_index] not in [None, '', 0]:
                data.append([
                    row[id_index],
                    row[article_index],
                    row[price_index],
                    row[weight_index],
                    row[product_index]
                ])

        # Копируем шаблон в новую директорию
        shutil.copy(price_and_weight_template, new_file_path)

        # Открываем скопированный файл
        wb_new = load_workbook(new_file_path)
        sheet_new = wb_new.active

        # Вставляем данные в новый файл, начиная с определенной строки
        # Начинаем со второй строки
        for row_index, data_row in enumerate(data, start=2):
            for col_index, value in enumerate(data_row):
                sheet_new.cell(
                    row=row_index, column=col_index + 1, value=value)

        # Сохраняем изменения в новом файле
        wb_new.save(new_file_path)


if __name__ == "__main__":
    settings_path = 'settings.json'  # Укажите путь к вашему файлу настроек
    price_handler = AvitoFilesHandler(settings_path)
    # price = price_handler.read_ostatki()
    # price = price_handler.update_price_from_products(price)
    # price = price_handler.apply_filters_to_price(price)
    price = price_handler.create_price_xlsx()
    # price_handler.create_price_file()
    # price_handler.create_all_avito_files()
    # price_handler.create_orders_files(["02", "03"], False)
    # price_handler.create_price_weigh_file()
    # price_handler.add_vba(8)
